import openpyxl
import re
from datetime import datetime


# ==========================================
# 勤務表解析：extract_duty_v2
# ==========================================
# 勤務表結構說明：
#   Row 3（index 2）：欄14起為時段標頭，格式「06\n|\n07」
#   Row 4（index 3）：通訊代號70值班，各時段填的是人員番號（A,B,D,01,02...）
#   Row 44~48（index 43~47）：番號對照表，每6格一組（番號, 職稱姓名 x5）
# ==========================================

def extract_duty_v2(file, current_hour: int) -> dict:
    """
    從勤務分配表 Excel 取出督導時段對應的值班人員與全員名冊。

    Parameters
    ----------
    file : UploadedFile（Streamlit）或 str（路徑）
        勤務分配表 .xlsx
    current_hour : int
        督導抵達的小時（0~23），用來對應勤務表時段欄

    Returns
    -------
    dict
        {
            'term'   : str,   # 單位名稱，如「中興派出所」
            'v_name' : str,   # 當時值班人員姓名
            'roster' : list,  # 全所人員姓名清單
        }
    """
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        # ---------- 1. 取單位名稱（Row1，欄0） ----------
        title_cell = str(all_rows[0][0]) if all_rows[0][0] else ''
        m_term = re.search(r'龍潭分局\s*([\u4e00-\u9fa5]+派出所|[\u4e00-\u9fa5]+隊)', title_cell)
        term = m_term.group(1) if m_term else '本所'

        # ---------- 2. 建立番號→姓名對照表（Row44~48） ----------
        code_map = {}
        for row in all_rows[43:48]:   # index 43~47 = Row44~48
            for grp in range(6):
                base = 1 + grp * 6
                if base + 1 >= len(row):
                    break
                code = row[base]
                name_cell = row[base + 1]
                if not code or code == '輪番番號':
                    continue
                if name_cell and isinstance(name_cell, str):
                    # 職稱+姓名，取最後2~4字中文作為姓名
                    names = re.findall(r'[\u4e00-\u9fa5]{2,4}', name_cell)
                    if names:
                        code_map[str(code).strip()] = names[-1]

        # ---------- 3. 找時段欄位索引 ----------
        time_headers = list(all_rows[2][13:])   # Row3，欄14起

        def find_col(hour):
            for i, h in enumerate(time_headers):
                if h is None:
                    continue
                h_str = str(h).split('\n')[0].strip()
                try:
                    if int(h_str) == hour:
                        return 13 + i
                except ValueError:
                    pass
            return None

        col = find_col(current_hour)

        # ---------- 4. 查值班番號（Row4） ----------
        v_code = ''
        if col is not None:
            raw = all_rows[3][col]   # Row4 = index 3
            if raw is not None:
                v_code = str(raw).strip()

        v_name = code_map.get(v_code, f'番號{v_code}')

        # ---------- 5. 全員名冊 ----------
        roster = list(code_map.values())

        return {
            'term'  : term,
            'v_name': v_name,
            'roster': roster,
        }

    except Exception as e:
        return {
            'term'  : '本所',
            'v_name': '（解析失敗）',
            'roster': [],
            '_error': str(e),
        }


# ==========================================
# 交接簿解析：extract_equip_v2
# ==========================================
# 交接簿結構說明：
#   Row 3（index 2）：欄位標頭（手槍、子彈、步槍...）
#   Row 4起：每4行一組（在所/出勤/送修/庫存）
#     欄0格式：「115-05-19 00:01\n交班人\n接班人」
#     欄2起：各裝備數量
# ==========================================

def extract_equip_v2(file) -> dict:
    """
    從值班人員交接登記簿 Excel 取出最新一筆交接紀錄，
    確認槍彈裝備數量均正常（送修欄為 0）。

    Parameters
    ----------
    file : UploadedFile（Streamlit）或 str（路徑）
        交接登記簿 .xlsx

    Returns
    -------
    dict
        {
            'ok'         : bool,   # 裝備是否全部正常（送修=0）
            'latest_time': str,    # 最新交接時間字串
            'from_officer': str,   # 交班人
            'to_officer'  : str,   # 接班人
            'summary'    : str,    # 人類可讀的裝備摘要
            'anomalies'  : list,   # 異常項目清單（送修>0的項目）
        }
    """
    # 欄位標頭（欄位順序固定）
    EQUIP_COLS = {
        2 : '手槍',
        3 : '手槍子彈',
        4 : '65式步槍',
        5 : '65式步槍子彈',
        6 : 'M16步槍',
        7 : 'M16步槍子彈',
        8 : '無線電',
        9 : '行動電腦',
        10: '防彈背心',
        11: '酒測器',
        12: '量知器',
        13: '電擊器',
        14: '警用機車',
        15: '巡邏車',
        16: '防彈頭盔',
    }

    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True))

        # 找所有交接組，格式為「在所」那行（每組第1行）
        groups = []
        seen_times = set()   # 避免同一交接時間被在所/出勤/送修/庫存4行都加入
        for i, row in enumerate(all_rows[3:], start=3):
            cell = row[0]
            if not cell or not isinstance(cell, str):
                continue
            if not re.search(r'\d{2}:\d{2}', cell):
                continue
            parts = cell.strip().split('\n')
            if len(parts) < 2:
                continue
            time_str = parts[0].strip()
            if time_str in seen_times:
                continue     # 同一交接時間只取第一次出現（在所行）
            seen_times.add(time_str)
            from_officer = parts[1].strip() if len(parts) > 1 else ''
            to_officer   = parts[2].strip() if len(parts) > 2 else ''
            m_time = re.search(r'\d{1,3}-\d{2}-\d{2} \d{2}:\d{2}', time_str)
            if m_time:
                groups.append({
                    'row_index'   : i,
                    'time_str'    : time_str,
                    'from_officer': from_officer,
                    'to_officer'  : to_officer,
                })

        if not groups:
            return {
                'ok': False, 'latest_time': '', 'from_officer': '',
                'to_officer': '', 'summary': '無交接紀錄', 'anomalies': [],
            }

        # 取最後一組（最新）
        latest = groups[-1]
        base_i = latest['row_index']

        # 取「送修」那行（第3行，offset=2）
        repair_row = all_rows[base_i + 2] if base_i + 2 < len(all_rows) else []
        # 取「在所」那行（offset=0）
        inservice_row = all_rows[base_i]

        anomalies = []
        for col_i, name in EQUIP_COLS.items():
            if col_i < len(repair_row):
                val = repair_row[col_i]
                if val and isinstance(val, (int, float)) and val > 0:
                    anomalies.append(f'{name}送修{int(val)}件')

        # 摘要：在所手槍+子彈
        gun_cnt    = inservice_row[2] if 2 < len(inservice_row) else '?'
        bullet_cnt = inservice_row[3] if 3 < len(inservice_row) else '?'
        summary = (
            f"在所手槍{gun_cnt}支、子彈{bullet_cnt}發，"
            f"{'裝備全部正常' if not anomalies else '異常：' + '、'.join(anomalies)}"
        )

        return {
            'ok'          : len(anomalies) == 0,
            'latest_time' : latest['time_str'],
            'from_officer': latest['from_officer'],
            'to_officer'  : latest['to_officer'],
            'summary'     : summary,
            'anomalies'   : anomalies,
        }

    except Exception as e:
        return {
            'ok': False, 'latest_time': '', 'from_officer': '',
            'to_officer': '', 'summary': f'解析失敗：{e}', 'anomalies': [],
        }


# ==========================================
# 快速測試（在本機直接執行此檔案時）
# ==========================================
if __name__ == '__main__':
    DUTY_FILE = '中興派出所1150519勤務分配表.xlsx'
    EQUIP_FILE = '桃園市政府警察局龍潭分局中興派出所值班人員交接登記簿.xlsx'

    print('=== extract_duty_v2（抵達時間=10時）===')
    dr = extract_duty_v2(DUTY_FILE, current_hour=10)
    print(dr)

    print('\n=== extract_equip_v2 ===')
    er = extract_equip_v2(EQUIP_FILE)
    print(er)
