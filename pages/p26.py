import streamlit as st
import pandas as pd
import re
import io
import os
import smtplib
import urllib.parse as _ul
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

# ==========================================
# 0. 輔助函式：發送信件至自己的信箱
# ==========================================
def send_file_email(file_bytes, file_name, unit_name, mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"):
    try:
        # 完全參照您的設定，讀取 user (sender) 與 password，並寄給自己
        sender = st.secrets["email"]["user"]
        pwd = st.secrets["email"]["password"]
        
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = sender
        
        msg["Subject"] = f"🚔 交通違規舉發績效結算表 - {unit_name}"
        
        body_text = (
            f"長官您好，\n\n"
            f"系統已自動產出最新結算的【交通違規舉發績效結算表與個人明細】。\n\n"
            f"結算單位：{unit_name}\n\n"
            "本信件由交通執法自動化分析引擎發送。"
        )
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        main_type, sub_type = mime_type.split('/') if '/' in mime_type else ("application", "octet-stream")
        part = MIMEBase(main_type, sub_type)
        part.set_payload(file_bytes.getvalue())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename*=UTF-8''{_ul.quote(file_name)}")
        msg.attach(part)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, pwd)
            server.sendmail(sender, sender, msg.as_string())
            
        return True, None
    except KeyError:
        return False, "系統找不到 Email 寄件設定。請確認 `secrets.toml` 中已設定 `[email]` 區塊，並包含 `user` 與 `password`。"
    except smtplib.SMTPAuthenticationError:
        return False, "Email 帳號或密碼驗證失敗！請確認使用的是「應用程式密碼」而非一般登入密碼。"
    except Exception as e:
        return False, str(e)

# ==========================================
# 1. 頁面基本設定與側邊欄
# ==========================================
st.set_page_config(page_title="舉發績效結算", page_icon="👮", layout="wide")

st.title("⚡ 員警交通違規舉發績效結算")
st.markdown("本頁面專門處理**員警個人績效配分**與**門檻結算**。匯出的報表將保留原始格式，若原始檔案缺少配分欄位，系統將自動安插並結算。算完後可下載或直接寄至您的信箱。")

st.sidebar.header("⚙️ 結算參數設定")
unit_type = st.sidebar.radio(
    "🏢 選擇單位與基準", 
    ["龍潭交通分隊 (基準800)", "一般單位 (基準400)"],
    index=0 
)
quota = 800 if "龍潭" in unit_type else 400
threshold_7x = quota * 7

st.sidebar.markdown("---")
st.sidebar.subheader("📂 步驟 1：上傳配分表")
db_file = st.sidebar.file_uploader("外部配分表 (如：檔案 B)", type=["xlsx"], key="db_file")

st.sidebar.subheader("📂 步驟 2：上傳原始舉發資料")
data_files = st.sidebar.file_uploader("批次選擇多個員警的半年期 Excel 檔案 (支援原始無配分欄位格式)", type=["xlsx", "xls"], accept_multiple_files=True, key="data_files")

# ==========================================
# 2. 核心結算邏輯
# ==========================================
if db_file and data_files:
    with st.spinner("🔄 正在讀取與比對資料..."):
        try:
            df_db = pd.read_excel(db_file)
            if '違規條款' not in df_db.columns:
                st.error("❌ 配分表缺少『違規條款』欄位！請檢查檔案格式。")
                st.stop()
        except Exception as e:
            st.error(f"❌ 讀取配分表失敗：{e}")
            st.stop()

        db_map = {}
        for _, row in df_db.iterrows():
            rule = str(row.get('違規條款', '')).strip()
            if rule:
                s = pd.to_numeric(row.get('攔舉配分', 0), errors='coerce')
                d = pd.to_numeric(row.get('逕舉配分', 0), errors='coerce')
                db_map[rule] = {
                    'stop': 0 if pd.isna(s) else int(s),
                    'dir': 0 if pd.isna(d) else int(d)
                }

        processed_sheets = []
        
        def extract_officer_name(df_head):
            for r_idx, row in df_head.iterrows():
                for c_idx, val in enumerate(row.values):
                    val_str = str(val).strip()
                    if "舉發員警" in val_str:
                        clean = re.sub(r'舉發員警[:：]?', '', val_str).strip()
                        if clean: return clean
                        if c_idx + 1 < len(row.values):
                            next_val = str(row.values[c_idx + 1]).strip()
                            if next_val and next_val.lower() != 'nan':
                                return next_val
            return ""

        for f in data_files:
            f.seek(0)
            try:
                xls = pd.ExcelFile(f)
                for sheet_name in xls.sheet_names:
                    raw_df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
                    raw_df = raw_df.astype('object')
                    
                    header_idx = -1
                    officer_name = extract_officer_name(raw_df.head(20))
                    if not officer_name:
                        officer_name = sheet_name.strip()
                    
                    for idx, row in raw_df.iterrows():
                        if "違規條款" in row.astype(str).str.replace(" ", "").values:
                            header_idx = idx
                            break
                    
                    if header_idx == -1:
                        continue

                    header_row = [str(x).strip().replace(" ", "") for x in raw_df.iloc[header_idx]]
                    col_rule = header_row.index("違規條款") if "違規條款" in header_row else -1
                    col_s_cnt = header_row.index("攔停數") if "攔停數" in header_row else -1
                    col_d_cnt = header_row.index("逕舉數") if "逕舉數" in header_row else -1
                    
                    col_s_score = header_row.index("攔舉配分") if "攔舉配分" in header_row else -1
                    col_d_score = header_row.index("逕舉配分") if "逕舉配分" in header_row else -1
                    col_subtotal = header_row.index("小計") if "小計" in header_row else -1
                    
                    if col_rule == -1 or col_s_cnt == -1 or col_d_cnt == -1:
                        continue

                    # 動態插入欄位
                    if col_s_score == -1:
                        idx_s_score = col_s_cnt + 1
                        raw_df.insert(idx_s_score, f'new_{idx_s_score}', None)
                        raw_df.iat[header_idx, idx_s_score] = "攔舉配分"
                        col_s_score = idx_s_score
                        
                        if col_d_cnt >= idx_s_score: col_d_cnt += 1
                        if col_subtotal >= idx_s_score: col_subtotal += 1
                        
                        idx_d_score = col_d_cnt + 1
                        raw_df.insert(idx_d_score, f'new_{idx_d_score}', None)
                        raw_df.iat[header_idx, idx_d_score] = "逕舉配分"
                        col_d_score = idx_d_score
                        
                        if col_subtotal >= idx_d_score: col_subtotal += 1
                        
                        idx_subtotal = idx_d_score + 1
                        raw_df.insert(idx_subtotal, f'new_{idx_subtotal}', None)
                        raw_df.iat[header_idx, idx_subtotal] = "小計"
                        col_subtotal = idx_subtotal
                        raw_df.columns = range(raw_df.shape[1])

                    grand_total = 0
                    yellow_cells = []
                    
                    for r in range(header_idx + 1, len(raw_df)):
                        rule = str(raw_df.iloc[r, col_rule]).strip()
                        if not rule or "合計" in rule or "製表" in rule or "舉發單張數" in rule or rule.lower() == 'nan':
                            continue
                            
                        def safe_int(val):
                            try: return int(float(str(val).replace(",", "")))
                            except: return 0
                                
                        stop_cnt = safe_int(raw_df.iloc[r, col_s_cnt])
                        dir_cnt = safe_int(raw_df.iloc[r, col_d_cnt])
                        
                        if rule in db_map:
                            s_score = db_map[rule]['stop']
                            d_score = db_map[rule]['dir']
                        else:
                            s_score = 0
                            d_score = 0
                            
                        raw_df.iat[r, col_s_score] = s_score if s_score != 0 else 0
                        raw_df.iat[r, col_d_score] = d_score if d_score != 0 else 0
                        
                        row_subtotal = (s_score * stop_cnt) + (d_score * dir_cnt)
                        if col_subtotal != -1:
                            raw_df.iat[r, col_subtotal] = row_subtotal
                        
                        if s_score == 0 and d_score == 0:
                            yellow_cells.append((r + 1, col_s_score + 1))
                            yellow_cells.append((r + 1, col_d_score + 1))
                            
                        grand_total += row_subtotal
                        
                    processed_sheets.append({
                        "officer": officer_name.replace(" ", ""),
                        "df": raw_df,
                        "yellow_cells": yellow_cells,
                        "grand_total": grand_total,
                        "col_d_score": col_d_score
                    })
                    
            except Exception as e:
                st.error(f"❌ 處理 {f.name} 時發生錯誤：{e}")

        # ==========================================
        # 3. 結算彙整與報表產出
        # ==========================================
        if processed_sheets:
            df_raw_summary = pd.DataFrame([{
                "員警姓名": s["officer"],
                "本期總分": s["grand_total"]
            } for s in processed_sheets])
            
            df_summary = df_raw_summary.groupby('員警姓名', as_index=False)['本期總分'].sum()
            df_summary['上半年分數'] = 5800  
            
            def calc_rem(score):
                if score >= threshold_7x: return score - threshold_7x
                return score % quota
                
            df_summary['上半年剩餘分數'] = df_summary['上半年分數'].apply(calc_rem)
            df_summary['最終總分'] = df_summary['本期總分'] + df_summary['上半年剩餘分數']

            st.success("✅ 所有原始報表已自動擴充欄位、結算並重建完畢！")
            st.subheader("📊 員警績效結算總表")
            st.dataframe(df_summary, use_container_width=True, hide_index=True)

            # --- 產出 Excel 檔案 ---
            output = io.BytesIO()
            wb = Workbook()
            
            ws_summary = wb.active
            ws_summary.title = "績效結算總表"
            ws_summary['A1'] = "交通違規舉發績效結算表"
            ws_summary['A1'].font = Font(size=14, bold=True)
            ws_summary['A2'] = f"結算單位：{unit_type}"
            
            header = list(df_summary.columns)
            header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
            for c_idx, title in enumerate(header, 1):
                cell = ws_summary.cell(row=4, column=c_idx, value=title)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = header_fill
                ws_summary.column_dimensions[get_column_letter(c_idx)].width = 16
                
            for r_idx, row_data in enumerate(df_summary.values, 5):
                for c_idx, val in enumerate(row_data, 1):
                    ws_summary.cell(row=r_idx, column=c_idx, value=val)

            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            sheet_name_counts = {}
            
            for s in processed_sheets:
                base_name = s["officer"][:25]
                if base_name not in sheet_name_counts:
                    sheet_name_counts[base_name] = 1
                    final_name = base_name
                else:
                    sheet_name_counts[base_name] += 1
                    final_name = f"{base_name}({sheet_name_counts[base_name]})"
                    
                ws_officer = wb.create_sheet(title=final_name)
                
                for r_idx, row in s["df"].iterrows():
                    cleaned_row = [val if pd.notna(val) else None for val in row.tolist()]
                    ws_officer.append(cleaned_row)
                    
                for r, c in s["yellow_cells"]:
                    ws_officer.cell(row=r, column=c).fill = yellow_fill
                    
                officer_summary = df_summary[df_summary['員警姓名'] == s['officer']].iloc[0]
                
                footer_start_row = ws_officer.max_row + 2
                write_col = s["col_d_score"] + 1 if s["col_d_score"] != -1 else 5
                
                def write_footer(row_offset, title, val, color="000000"):
                    c_title = ws_officer.cell(row=footer_start_row + row_offset, column=write_col - 1, value=title)
                    c_title.font = Font(bold=True)
                    c_val = ws_officer.cell(row=footer_start_row + row_offset, column=write_col, value=val)
                    c_val.font = Font(bold=True, color=color)
                
                write_footer(0, "本期分數：", int(officer_summary["本期總分"]), "0000FF")
                write_footer(1, "上半年分數：", int(officer_summary["上半年分數"]), "0000FF")
                write_footer(2, "上半年剩餘分數：", int(officer_summary["上半年剩餘分數"]), "0000FF")
                write_footer(3, "總分：", int(officer_summary["最終總分"]), "FF0000")
                
                for col_idx in range(1, len(s["df"].columns) + 1):
                    ws_officer.column_dimensions[get_column_letter(col_idx)].width = 12

            wb.save(output)
            
            # ==========================================
            # 4. 檔案下載與信件寄送區塊
            # ==========================================
            st.divider()
            col1, col2 = st.columns(2)
            
            excel_filename = f"舉發績效結算與個人明細表_{unit_type[:2]}.xlsx"
            
            with col1:
                st.markdown("### 📥 下載報表")
                st.download_button(
                    label="下載完整報表 (含總表與明細)",
                    data=output.getvalue(),
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

            with col2:
                st.markdown("### 📧 寄送報表至我的信箱")
                
                if st.button("🚀 一鍵寄送報表", use_container_width=True):
                    with st.spinner("信件發送中，請稍候…"):
                        ok, mail_err = send_file_email(
                            file_bytes=output, 
                            file_name=excel_filename, 
                            unit_name=unit_type
                        )
                        if ok:
                            st.success("✅ 信件發送成功！報表已隨信夾帶至您的信箱。")
                            st.balloons()
                        else:
                            st.error(f"❌ 發信失敗: {mail_err}")
