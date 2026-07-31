import streamlit as st
import pandas as pd
import re
import io
import os
import smtplib
import urllib.parse as _ul
import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import gspread
from google.oauth2.service_account import Credentials

# ==========================================
# 💡 PDF 產出相關套件 (ReportLab)
# ==========================================
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# 設定標楷體字型
font_path = 'kaiu.ttf'
if os.path.exists(font_path):
    pdfmetrics.registerFont(TTFont('KaiTi', font_path))
    FONT_NAME = 'KaiTi'
else:
    fallback_path = '/usr/share/fonts/truetype/droid/DroidSansFallbackFull.ttf'
    if os.path.exists(fallback_path):
        pdfmetrics.registerFont(TTFont('KaiTi', fallback_path))
        FONT_NAME = 'KaiTi'
    else:
        FONT_NAME = 'Helvetica'

# ==========================================
# 💡 匯入系統原本的側邊欄設定
# ==========================================
try:
    from menu import show_sidebar
except ImportError:
    def show_sidebar():
        pass

# ==========================================
# 💡 系統後台設定
# ==========================================
TARGET_GSHEET_URL = "https://docs.google.com/spreadsheets/d/1HaFu5PZkFDUg7WZGV9khyQ0itdGXhXUakP4_BClFTUg/edit"

def get_gspread_client():
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive"
    ]
    if "gcp_service_account" in st.secrets:
        creds_dict = dict(st.secrets["gcp_service_account"])
    elif "connections" in st.secrets and "gsheets" in st.secrets["connections"]:
        creds_dict = dict(st.secrets["connections"]["gsheets"])
    else:
        return None
    creds = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    return gspread.authorize(creds)

# ==========================================
# 0. 輔助函式：發送 Email
# ==========================================
def send_multiple_files_email(file_buffers_dict):
    try:
        sender = st.secrets["email"]["user"]
        pwd = st.secrets["email"]["password"]
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = sender
        msg["Subject"] = "🚔 交通執法重點工作舉發績效結算表 (批次產出)"
        
        body_text = "長官您好，\n\n系統已自動產出最新結算的【交通執法重點工作舉發績效結算表與個人明細】。\n\n本次共產出以下單位報表，詳見附件：\n"
        for fname in file_buffers_dict.keys():
            body_text += f"- {fname}\n"
        body_text += "\n本信件由交通執法自動化分析引擎發送。"
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        for fname, buffer in file_buffers_dict.items():
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(buffer.getvalue())
            encoders.encode_base64(part)
            part.add_header("Content-Disposition", f"attachment; filename*=UTF-8''{_ul.quote(fname)}")
            msg.attach(part)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, pwd)
            server.sendmail(sender, sender, msg.as_string())
        return True, None
    except Exception as e:
        return False, str(e)

def send_single_file_email(file_bytes, file_name, mime_type="application/pdf"):
    try:
        sender = st.secrets["email"]["user"]
        pwd = st.secrets["email"]["password"]
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = sender
        msg["Subject"] = f"🚔 交通執法重點工作自動產出結果_{file_name}"
        
        body_text = f"長官您好，\n\n系統已自動產出相關檔案，附件為您要求的 {file_name}。\n\n本信件由交通執法自動化分析引擎發送。"
        msg.attach(MIMEText(body_text, "plain", "utf-8"))

        main_type, sub_type = mime_type.split('/') if '/' in mime_type else ("application", "octet-stream")
        part = MIMEBase(main_type, sub_type)
        part.set_payload(file_bytes.getvalue())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename*=UTF-8''{_ul.quote(file_name)}")
        msg.attach(part)

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender, pwd)
            server.sendmail(sender, sender, msg.as_string())
        return True, None
    except Exception as e:
        return False, str(e)

# ==========================================
# 0. 輔助函式：自動調整 Excel 欄寬與字串長度計算
# ==========================================
def get_display_width(text):
    if not text: return 0
    width = 0
    for char in str(text):
        if ord(char) > 128: width += 2.1
        else: width += 1.1
    return width

def autofit_columns(ws, min_row=1, max_width=60):
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.row < min_row: continue
            if cell.value is not None:
                w = get_display_width(cell.value)
                if w > max_length: max_length = w
        if max_length > 0:
            ws.column_dimensions[col_letter].width = min(max_length + 2, max_width)

# ==========================================
# 💡 核心 PDF 產出邏輯
# ==========================================
def generate_traffic_enforcement_pdf(target_roc_year, period_text, months_text):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, right_margin=30, left_margin=30, top_margin=15, bottom_margin=15)
    story = []
    
    title_style = ParagraphStyle('TitleStyle', fontName=FONT_NAME, fontSize=17, leading=22, alignment=1)
    header_style = ParagraphStyle('HeaderStyle', fontName=FONT_NAME, fontSize=13, leading=16, alignment=1)
    body_style = ParagraphStyle('BodyStyle', fontName=FONT_NAME, fontSize=13, leading=18, alignment=4)
    sign_style = ParagraphStyle('SignStyle', fontName=FONT_NAME, fontSize=13, leading=18, alignment=0)
    
    char_w = 13
    body_l1 = ParagraphStyle('BodyL1', parent=body_style, leftIndent=char_w*2, firstLineIndent=-char_w*2)
    body_step = ParagraphStyle('BodyStep', parent=body_style, leftIndent=char_w*2, firstLineIndent=0)

    now = datetime.datetime.now()
    current_date_str = f"{now.year - 1911}年{now.month}月{now.day}日"
    deadline_date = now + datetime.timedelta(days=7)
    deadline_str = f"{deadline_date.year - 1911}年{deadline_date.month}月{deadline_date.day}日"

    units = ["聖亭所", "龍潭所", "中興所", "石門所", "高平所", "龍潭交通分隊", "三和所", "警備隊"]

    for idx, unit_name in enumerate(units):
        story.append(Paragraph("桃園市政府警察局龍潭分局交通組交辦單", title_style))
        story.append(Spacer(1, 8))
        
        notice_paragraphs = [
            Paragraph(f"一、為辦理本分局{target_roc_year}年{period_text}({months_text})執行「交通執法重點工作」出力人員敘獎案，請至「人事資源整合管理系統」登錄敘獎人員獎勵資料。", body_l1),
            Paragraph("二、獎勵規則：", body_l1),
            Paragraph("按「警察機關交通執法獎懲作業規定」三、「警察人員交通執法之獎勵，行政警察每四百分嘉獎一次，交通警察每八百分嘉獎一次，每半年總獎勵額度最高以嘉獎七次為限；下半年總獎勵尚未達嘉獎七次者，未達獎勵基準之剩餘分數，行政警察剩餘分數超過二百分嘉獎一次，交通警察剩餘分數超過四百分嘉獎一次；上半年未敘至嘉獎七次者，其未達獎勵基準之剩餘分數，得併下半年合計」。", body_step),
            Paragraph("三、請至網路硬碟/交通組/巡官郭勝隆的資料夾/員警開單績效統計表查閱「總分」，照上述獎勵規則，於「人事資訊整合管理系統」登錄獎懲勵資料完畢後，毋庸附件，將交辦單逕送本組。", body_l1),
            Paragraph("四、登錄獎勵資料的流程：", body_l1),
            Paragraph(f"新增>主旨事由:{target_roc_year}年{period_text}({months_text})執行「交通執法重點工作」出力人員敘獎案>受理單位代碼:交通組>受理人員代碼:郭勝隆>再按新增>加入獎懲人員>獎懲事由:{target_roc_year}年{period_text}執行「交通執法重點工作」達幾分。", body_step),
            Paragraph(f"五、若超過2次嘉獎以上，獎懲事由則輸入{target_roc_year}年{period_text}執行「交通執法重點工作」達幾分-1、-2、-3(以此類推)。", body_l1),
            Paragraph("六、請不用寫辛勞得力及備極辛勞，系統會自動帶入。", body_l1),
            Spacer(1, 5),
            # 💡 已刪除「連同原件具報」
            Paragraph(f"辦理期限：{deadline_str}前辦理完畢。", body_style)
        ]
        
        sign_content = (
            '承辦人：<br/><br/><br/><br/>單位主管：<br/><br/><br/>'
            '<font color="white">白白白白白白白</font>年'
            '<font color="white">白白白</font>月'
            '<font color="white">白白白</font>日'
        )

        data = [
            [
                Paragraph("受文者", header_style), Paragraph(unit_name, header_style), 
                Paragraph("交辦日期", header_style), Paragraph(current_date_str, header_style), 
                Paragraph("承辦人", header_style), Paragraph("", header_style), 
                Paragraph("單位主管", header_style), Paragraph("組長楊孟竟", header_style)
            ],
            [
                Paragraph("交<br/><br/>辦<br/><br/>事<br/><br/>由", header_style), 
                notice_paragraphs, '', '', '', '', '', ''
            ],
            [
                Paragraph("承 辦 內<br/>容", header_style), 
                Paragraph(sign_content, sign_style), '', '', '', '', '', ''
            ]
        ]
        
        t = Table(data, colWidths=[56, 45, 66, 91, 45, 102, 56, 74])
        t.setStyle(TableStyle([
            ('GRID', (0,0), (-1,-1), 1, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('SPAN', (1,1), (7,1)),
            ('SPAN', (1,2), (7,2)),
            ('TOPPADDING', (0,0), (-1,-1), 4),
            ('BOTTOMPADDING', (0,0), (-1,-1), 4),
            ('LEFTPADDING', (0,0), (-1,-1), 2),
            ('RIGHTPADDING', (0,0), (-1,-1), 2),
        ]))
        
        story.append(t)
        if idx < len(units) - 1:
            story.append(PageBreak())

    doc.build(story)
    buffer.seek(0)
    return buffer

# ==========================================
# 主程式執行區塊
# ==========================================
def main():
    st.set_page_config(page_title="舉發績效結算", page_icon="👮", layout="wide")
    show_sidebar()

    st.title("⚡ 員警交通執法重點工作舉發績效結算")
    st.markdown("本頁面專門處理**員警個人績效配分**與**門檻結算**。系統具備雙向雲端同步與**智慧修復引擎**，並支援全年度期程結轉。")

    # ==========================================
    # 💡 共用的結算期程與年度設定
    # ==========================================
    st.markdown("### 📅 步驟 1：選擇結算年度與期程")
    col_y, col_p = st.columns(2)
    with col_y:
        now = datetime.datetime.now()
        default_roc_year = now.year - 1911
        if now.month <= 2:
            default_roc_year -= 1 # 1、2月預設結算前一年
        target_roc_year = st.number_input("結算年度 (民國年)", min_value=110, value=default_roc_year, step=1)
        
    with col_p:
        default_period_idx = 1 if now.month <= 2 else 0 
        period = st.radio("結算期程", ["上半年", "下半年"], index=default_period_idx, horizontal=True)
        
    months_text = "1至6月" if period == "上半年" else "7至12月"
    st.markdown("---")

    # 分離功能為兩個 Tabs
    tab_pdf, tab_excel = st.tabs(["📄 1. 各單位交辦單 PDF 產生器", "📊 2. 原始舉發資料結算與匯出"])

    # ==========================================
    # TAB 1: 交辦單 PDF 產生器
    # ==========================================
    with tab_pdf:
        st.subheader("📋 桃市警龍潭分局交通組交辦單 (全單位一次匯出)")
        st.write("已導入動態期程辨識，交辦內容的年份與期程將完全自動同步上方設定，辦理期限預設為今日的後 7 天。")
        
        pdf_data = generate_traffic_enforcement_pdf(target_roc_year, period, months_text)
        pdf_file_name = f"{target_roc_year}年{period}交通執法重點工作交辦單_全單位.pdf"
        
        col_pdf1, col_pdf2 = st.columns(2)
        with col_pdf1:
            st.download_button(
                label="📥 下載【全單位】交辦單 (PDF)",
                data=pdf_data,
                file_name=pdf_file_name,
                mime="application/pdf",
                use_container_width=True
            )
        with col_pdf2:
            if st.button("📧 將此 PDF 一鍵寄至我的信箱", use_container_width=True):
                with st.spinner("信件發送中，請稍候…"):
                    ok, mail_err = send_single_file_email(pdf_data, pdf_file_name, mime_type="application/pdf")
                    if ok:
                        st.success("✅ 信件發送成功！交辦單 PDF 已夾帶至您的信箱。")
                    else:
                        st.error(f"❌ 發信失敗: {mail_err}")

    # ==========================================
    # TAB 2: 原始舉發資料結算
    # ==========================================
    with tab_excel:
        db_map = {}
        df_db = None
        sheet_id = TARGET_GSHEET_URL.split("/d/")[1].split("/")[0]
        xlsx_export_url = f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

        try:
            df_db = pd.read_excel(xlsx_export_url, sheet_name="配分表")
            connection_status = True
        except Exception as e:
            connection_status = False
            error_msg = e

        incomplete_rules = [] 
        if connection_status:
            if '違規條款' not in df_db.columns:
                st.error("❌ 雲端配分表缺少『違規條款』欄位！")
                st.stop()
            if '違規事實' not in df_db.columns: df_db['違規事實'] = ""
            if '類別' not in df_db.columns: df_db['類別'] = ""
            if '取締項目' not in df_db.columns: df_db['取締項目'] = ""

            for row_idx, row in df_db.iterrows():
                rule = str(row.get('違規條款', '')).strip()
                if rule:
                    s = pd.to_numeric(row.get('攔舉配分', 0), errors='coerce')
                    d = pd.to_numeric(row.get('逕舉配分', 0), errors='coerce')
                    fact = str(row.get('違規事實', '')).strip()
                    cat = str(row.get('類別', '')).strip()
                    item = str(row.get('取締項目', '')).strip()
                    
                    if fact.lower() == 'nan': fact = ""
                    if cat.lower() == 'nan': cat = ""
                    if item.lower() == 'nan': item = ""
                        
                    db_map[rule] = {
                        'stop': 0 if pd.isna(s) else int(s),
                        'dir': 0 if pd.isna(d) else int(d),
                        'fact': fact,
                        'category': cat,
                        'item': item,
                        'gsheet_row': row_idx + 2 
                    }
                    
                    if not cat and not item:
                        incomplete_rules.append(rule)

        def find_closest_reference(new_rule_str, existing_db):
            best_match = None
            max_prefix_len = 0
            for old_rule, data in existing_db.items():
                match_len = 0
                for c1, c2 in zip(new_rule_str, old_rule):
                    if c1 == c2: match_len += 1
                    else: break
                if match_len > max_prefix_len and (data['category'] or data['item']):
                    max_prefix_len = match_len
                    best_match = data
            return best_match

        with st.expander("⚙️ 系統運行狀態與結算基準 (點擊展開)", expanded=bool(incomplete_rules)):
            if connection_status:
                st.success("🟢 雲端配分資料庫連線正常")
            else:
                st.error(f"🔴 雲端資料庫連線失敗：{error_msg}")
                st.stop()
                
            st.info("💡 **自動結算基準**\n- **交通分隊**：800 分\n- **其他單位**：400 分\n*(獎勵上限 7 倍，年度結算時下半年之殘餘分數將不予保留)*")

            if incomplete_rules:
                st.warning(f"⚠️ 偵測到雲端資料庫中有 **{len(incomplete_rules)}** 筆過往新增的條款缺乏「類別」與「取締項目」。")
                if st.button("🛠️ 一鍵智慧修復資料庫", use_container_width=True):
                    with st.spinner("☁️ 正在透過智慧參照引擎修復雲端資料..."):
                        client = get_gspread_client()
                        if not client:
                            st.error("❌ 找不到 GCP 憑證，無法寫入雲端。")
                        else:
                            try:
                                worksheet = client.open_by_key(sheet_id).worksheet("配分表")
                                updates = []
                                for rule in incomplete_rules:
                                    ref_data = find_closest_reference(rule, db_map)
                                    if ref_data:
                                        target_row = db_map[rule]['gsheet_row']
                                        new_cat = ref_data['category']
                                        new_item = ref_data['item']
                                        updates.append({'range': f'E{target_row}', 'values': [[new_cat]]})
                                        updates.append({'range': f'F{target_row}', 'values': [[new_item]]})
                                        db_map[rule]['category'] = new_cat
                                        db_map[rule]['item'] = new_item
                                if updates:
                                    worksheet.batch_update(updates) 
                                    st.success("✅ 修復完成！已成功將參照類別補上。")
                                    st.rerun() 
                                else:
                                    st.info("無可參照的相近條款。")
                            except Exception as e:
                                st.error(f"❌ 修復失敗：{e}")

        st.subheader("📂 步驟 2：資料上傳")
        
        h1_remainders = {}
        if period == "下半年":
            st.info("💡 下半年結算需加入前期的保留分數。請上傳前次產出的**【上半年績效結算總表】**(可多選)。")
            h1_files = st.file_uploader("上傳上半年績效總表 (Excel)", type=["xlsx", "xls"], accept_multiple_files=True, key="h1_files")
            
            if h1_files:
                for f in h1_files:
                    try:
                        df_h1 = pd.read_excel(f, sheet_name="績效結算總表", header=3)
                        if "員警姓名" in df_h1.columns and "上半年剩餘分數" in df_h1.columns:
                            for _, r in df_h1.iterrows():
                                name = str(r["員警姓名"]).strip()
                                rem = pd.to_numeric(r["上半年剩餘分數"], errors='coerce')
                                if name and pd.notna(rem):
                                    h1_remainders[name] = int(rem)
                    except Exception as e:
                        st.warning(f"⚠️ 無法讀取檔案 {f.name} 的餘數資料，請確認是否為系統產出的總表。")

        data_files = st.file_uploader(f"批次上傳各單位的【{period}原始舉發 Excel 報表】", type=["xlsx", "xls"], accept_multiple_files=True, key="data_files")

        if data_files:
            missing_rules = {} 
            for f in data_files:
                f.seek(0)
                try:
                    xls = pd.ExcelFile(f)
                    for sheet_name in xls.sheet_names:
                        raw_df = pd.read_excel(xls, sheet_name=sheet_name, header=None).astype('object')
                        header_idx = -1
                        for idx, row in raw_df.iterrows():
                            if "違規條款" in row.astype(str).str.replace(" ", "").values:
                                header_idx = idx
                                break
                        if header_idx == -1: continue

                        header_row = [str(x).strip().replace(" ", "") for x in raw_df.iloc[header_idx]]
                        if "違規條款" not in header_row: continue
                        
                        col_rule = header_row.index("違規條款")
                        col_fact = next((i for i, c in enumerate(header_row) if "違規事實" in c), -1)
                        col_cat = next((i for i, c in enumerate(header_row) if "類別" in c), -1)
                        col_item = next((i for i, c in enumerate(header_row) if "取締項目" in c), -1)

                        for r in range(header_idx + 1, len(raw_df)):
                            rule = str(raw_df.iloc[r, col_rule]).strip()
                            if not rule or "合計" in rule or "製表" in rule or "舉發單張數" in rule or rule.lower() == 'nan':
                                continue
                                
                            if rule not in db_map:
                                if rule not in missing_rules:
                                    missing_rules[rule] = {"fact": "", "category": "", "item": ""}
                                    
                                if not missing_rules[rule]["fact"] and col_fact != -1:
                                    val = str(raw_df.iloc[r, col_fact]).strip()
                                    if val.lower() != 'nan': missing_rules[rule]["fact"] = val
                                    
                                if not missing_rules[rule]["category"] and col_cat != -1:
                                    val = str(raw_df.iloc[r, col_cat]).strip()
                                    if val.lower() != 'nan': missing_rules[rule]["category"] = val
                                    
                                if not missing_rules[rule]["item"] and col_item != -1:
                                    val = str(raw_df.iloc[r, col_item]).strip()
                                    if val.lower() != 'nan': missing_rules[rule]["item"] = val
                except Exception:
                    pass

            for rule_key, data in missing_rules.items():
                if not data["category"] and not data["item"]:
                    ref_data = find_closest_reference(rule_key, db_map)
                    if ref_data:
                        data["category"] = ref_data["category"]
                        data["item"] = ref_data["item"]

            if missing_rules:
                st.warning(f"⚠️ 掃描完畢！發現 **{len(missing_rules)}** 筆新條款。系統已根據雲端資料庫智慧帶入相近類別，請確認配分：")
                
                missing_df = pd.DataFrame({
                    "違規條款": list(missing_rules.keys()),
                    "違規事實": [v["fact"] for v in missing_rules.values()],
                    "類別": [v["category"] for v in missing_rules.values()],
                    "取締項目": [v["item"] for v in missing_rules.values()],
                    "攔舉配分": 0,
                    "逕舉配分": 0
                })
                
                edited_missing = st.data_editor(
                    missing_df,
                    column_config={
                        "違規條款": st.column_config.TextColumn("違規條款 (未設定)", disabled=True),
                        "違規事實": st.column_config.TextColumn("違規事實", disabled=False),
                        "類別": st.column_config.TextColumn("類別 ✍️", disabled=False),
                        "取締項目": st.column_config.TextColumn("取締項目 ✍️", disabled=False),
                        "攔舉配分": st.column_config.NumberColumn("攔舉配分 ✍️", min_value=0, required=True),
                        "逕舉配分": st.column_config.NumberColumn("逕舉配分 ✍️", min_value=0, required=True),
                    },
                    hide_index=True,
                    use_container_width=True,
                    key="missing_rules_editor"
                )
                
                if not st.button("✅ 我已確認完畢，同步寫回雲端並開始結算", type="primary", use_container_width=True):
                    st.stop() 
                else:
                    with st.spinner("☁️ 正在將新條款「批次」同步寫回 Google 試算表..."):
                        client = get_gspread_client()
                        if not client:
                            st.error("❌ 找不到對應的 GCP 憑證設定。")
                            st.stop()
                        try:
                            worksheet = client.open_by_key(sheet_id).worksheet("配分表")
                            new_rows_data = []
                            for _, row in edited_missing.iterrows():
                                rule_val = str(row["違規條款"])
                                s_val = int(row["攔舉配分"])
                                d_val = int(row["逕舉配分"])
                                fact_val = str(row["違規事實"]) if pd.notna(row["違規事實"]) else ""
                                cat_val = str(row["類別"]) if pd.notna(row["類別"]) else ""
                                item_val = str(row["取締項目"]) if pd.notna(row["取締項目"]) else ""
                                
                                new_rows_data.append([rule_val, s_val, d_val, fact_val, cat_val, item_val])
                                db_map[rule_val] = {'stop': s_val, 'dir': d_val}
                                
                            if new_rows_data:
                                worksheet.append_rows(new_rows_data)
                            st.success("✅ 新條款與對應資訊已「整批」成功寫回 Google 試算表最下方！")
                        except Exception as write_err:
                            st.error(f"❌ 寫入雲端失敗，詳細錯誤：{write_err}")
                            st.stop()

            with st.spinner("🔄 正在讀取並結算各單位資料，請稍候..."):
                for f in data_files:
                    f.seek(0)
                    
                def extract_officer_name(df_head):
                    for r_idx, row in df_head.iterrows():
                        for c_idx, val in enumerate(row.values):
                            val_str = str(val).strip()
                            if "舉發員警" in val_str:
                                clean = re.sub(r'舉發員警[:：]?', '', val_str).strip()
                                if clean: return clean
                                if c_idx + 1 < len(row.values):
                                    next_val = str(row.values[c_idx + 1]).strip()
                                    if next_val and next_val.lower() != 'nan': return next_val
                    return ""

                def extract_unit_name(df_head):
                    for r_idx, row in df_head.iterrows():
                        for c_idx, val in enumerate(row.values):
                            val_str = str(val).strip()
                            if "舉發單位" in val_str:
                                clean = re.sub(r'舉發單位[:：]?', '', val_str).strip()
                                if clean: return clean
                                if c_idx + 1 < len(row.values):
                                    next_val = str(row.values[c_idx + 1]).strip()
                                    if next_val and next_val.lower() != 'nan': return next_val
                    return ""

                unit_collected_data = {}

                for f in data_files:
                    try:
                        xls = pd.ExcelFile(f)
                        for sheet_name in xls.sheet_names:
                            raw_df = pd.read_excel(xls, sheet_name=sheet_name, header=None).astype('object')
                            header_idx = -1
                            officer_name = extract_officer_name(raw_df.head(20))
                            if not officer_name: officer_name = sheet_name.strip()
                                
                            detected_unit = extract_unit_name(raw_df.head(20))
                            if not detected_unit: detected_unit = re.sub(r'\.[a-zA-Z0-9]+$', '', f.name) 
                            
                            for idx, row in raw_df.iterrows():
                                if "違規條款" in row.astype(str).str.replace(" ", "").values:
                                    header_idx = idx
                                    break
                            if header_idx == -1: continue
                            
                            print_date_val = ""
                            issue_date_val = ""
                            unit_val = ""
                            officer_val = ""
                            
                            for r_search in range(header_idx):
                                for c_search in range(len(raw_df.columns)):
                                    val = raw_df.iat[r_search, c_search]
                                    if pd.notna(val) and str(val).strip() != "":
                                        val_str = str(val).strip()
                                        val_no_space = val_str.replace(" ", "")
                                        if "列印日期" in val_no_space: print_date_val = val_str
                                        elif "開單日期" in val_no_space: issue_date_val = val_str
                                        elif "舉發單位" in val_no_space: unit_val = val_str
                                        elif "舉發員警" in val_no_space: officer_val = val_str

                            raw_df = raw_df.iloc[header_idx:].reset_index(drop=True)
                            header_idx = 0 
                            
                            header_row_temp = [str(x).strip().replace(" ", "") for x in raw_df.iloc[header_idx]]
                            cols_to_keep = [c for c, val in enumerate(header_row_temp) if val not in ["nan", "None", ""]]
                            
                            raw_df = raw_df.iloc[:, cols_to_keep]
                            raw_df.columns = range(raw_df.shape[1])

                            header_row = [str(x).strip().replace(" ", "") for x in raw_df.iloc[header_idx]]
                            col_rule = header_row.index("違規條款") if "違規條款" in header_row else -1
                            col_s_cnt = header_row.index("攔停數") if "攔停數" in header_row else -1
                            col_d_cnt = header_row.index("逕舉數") if "逕舉數" in header_row else -1
                            
                            col_s_score = header_row.index("攔舉配分") if "攔舉配分" in header_row else -1
                            col_d_score = header_row.index("逕舉配分") if "逕舉配分" in header_row else -1
                            col_subtotal = header_row.index("小計") if "小計" in header_row else -1
                            
                            if col_rule == -1 or col_s_cnt == -1 or col_d_cnt == -1: continue

                            if col_s_score == -1:
                                idx_s_score = col_s_cnt + 1
                                raw_df.insert(idx_s_score, f'new_{idx_s_score}', None)
                                raw_df.iat[header_idx, idx_s_score] = "攔舉配分"
                                col_s_score = idx_s_score
                                
                                if col_d_cnt >= idx_s_score: col_d_cnt += 1
                                if col_subtotal >= idx_s_score: col_subtotal += 1
                                
                                idx_d_score = col_d_cnt + 1
                                raw_df.insert(idx_d_score, f'new_{idx_d_score}', None)
                                raw_df.iat[header_idx, idx_d_score] = "逕舉配分"
                                col_d_score = idx_d_score
                                
                                if col_subtotal >= idx_d_score: col_subtotal += 1
                                
                                idx_subtotal = idx_d_score + 1
                                raw_df.insert(idx_subtotal, f'new_{idx_subtotal}', None)
                                raw_df.iat[header_idx, idx_subtotal] = "小計"
                                col_subtotal = idx_subtotal
                                raw_df.columns = range(raw_df.shape[1])

                            grand_total = 0
                            yellow_cells = []
                            
                            for r in range(header_idx + 1, len(raw_df)):
                                rule = str(raw_df.iloc[r, col_rule]).strip()
                                if not rule or "合計" in rule or "製表" in rule or "舉發單張數" in rule or rule.lower() == 'nan':
                                    continue
                                    
                                def safe_int(val):
                                    try: return int(float(str(val).replace(",", "")))
                                    except: return 0
                                        
                                stop_cnt = safe_int(raw_df.iloc[r, col_s_cnt])
                                dir_cnt = safe_int(raw_df.iloc[r, col_d_cnt])
                                
                                if rule in db_map:
                                    s_score = db_map[rule]['stop']
                                    d_score = db_map[rule]['dir']
                                else:
                                    s_score = 0
                                    d_score = 0
                                    
                                raw_df.iat[r, col_s_score] = s_score if s_score != 0 else 0
                                raw_df.iat[r, col_d_score] = d_score if d_score != 0 else 0
                                
                                row_subtotal = (s_score * stop_cnt) + (d_score * dir_cnt)
                                if col_subtotal != -1:
                                    raw_df.iat[r, col_subtotal] = row_subtotal
                                
                                if s_score == 0 and d_score == 0:
                                    yellow_cells.append((r, col_s_score))
                                    yellow_cells.append((r, col_d_score))
                                    
                                grand_total += row_subtotal
                                
                            sheet_data = {
                                "officer": officer_name.replace(" ", ""),
                                "df": raw_df,
                                "yellow_cells": yellow_cells,
                                "grand_total": grand_total,
                                "col_d_score": col_d_score,
                                "print_date": print_date_val,
                                "issue_date": issue_date_val,
                                "unit_val": unit_val,
                                "officer_val": officer_val
                            }
                            
                            if detected_unit not in unit_collected_data:
                                is_800 = "交通分隊" in detected_unit
                                quota_val = 800 if is_800 else 400
                                unit_collected_data[detected_unit] = {
                                    "processed_sheets": [],
                                    "quota": quota_val,
                                    "threshold_7x": quota_val * 7,
                                    "unit_type_label": f"{detected_unit} (基準 {quota_val} 分)"
                                }
                            unit_collected_data[detected_unit]["processed_sheets"].append(sheet_data)
                            
                    except Exception as e:
                        st.error(f"❌ 處理檔案 {f.name} 時發生錯誤：{e}")
                        continue

                all_summaries = {}
                all_output_buffers = {}

                for unit_name, unit_info in unit_collected_data.items():
                    processed_sheets = unit_info["processed_sheets"]
                    quota = unit_info["quota"]
                    threshold_7x = unit_info["threshold_7x"]
                    unit_type_label = unit_info["unit_type_label"]

                    if processed_sheets:
                        if period == "上半年":
                            df_raw_summary = pd.DataFrame([{
                                "員警姓名": s["officer"],
                                "上半年總分": s["grand_total"]
                            } for s in processed_sheets])
                            df_summary = df_raw_summary.groupby('員警姓名', as_index=False)['上半年總分'].sum()
                            
                            def calc_rem(score):
                                if score >= threshold_7x: return score - threshold_7x
                                return score % quota
                            df_summary['上半年剩餘分數'] = df_summary['上半年總分'].apply(calc_rem)
                            
                        else: 
                            df_raw_summary = pd.DataFrame([{
                                "員警姓名": s["officer"],
                                "本期原始分數": s["grand_total"]
                            } for s in processed_sheets])
                            df_summary = df_raw_summary.groupby('員警姓名', as_index=False)['本期原始分數'].sum()
                            df_summary['上半年結轉餘數'] = df_summary['員警姓名'].map(h1_remainders).fillna(0).astype(int)
                            df_summary['下半年總分'] = df_summary['本期原始分數'] + df_summary['上半年結轉餘數']

                        all_summaries[unit_name] = df_summary

                        output = io.BytesIO()
                        wb = Workbook()
                        
                        ws_summary = wb.active
                        ws_summary.title = "績效結算總表"
                        ws_summary['A1'] = f"交通執法重點工作舉發績效結算表 ({period})"
                        ws_summary['A1'].font = Font(size=14, bold=True)
                        ws_summary['A2'] = f"結算單位：{unit_type_label}"
                        
                        ws_summary.page_setup.orientation = "portrait"
                        ws_summary.page_setup.fitToWidth = 1
                        ws_summary.page_setup.fitToHeight = 0
                        ws_summary.sheet_properties.pageSetUpPr.fitToPage = True
                        
                        header = list(df_summary.columns)
                        header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
                        for c_idx, title in enumerate(header, 1):
                            cell = ws_summary.cell(row=4, column=c_idx, value=title)
                            cell.font = Font(bold=True, color="FFFFFF")
                            cell.fill = header_fill
                            
                        for r_idx, row_data in enumerate(df_summary.values, 5):
                            for c_idx, val in enumerate(row_data, 1):
                                ws_summary.cell(row=r_idx, column=c_idx, value=val)

                        autofit_columns(ws_summary, min_row=4, max_width=40)

                        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                        sheet_name_counts = {}
                        
                        for s in processed_sheets:
                            base_name = s["officer"][:25]
                            if base_name not in sheet_name_counts:
                                sheet_name_counts[base_name] = 1
                                final_name = base_name
                            else:
                                sheet_name_counts[base_name] += 1
                                final_name = f"{base_name}({sheet_name_counts[base_name]})"
                                
                            ws_officer = wb.create_sheet(title=final_name)
                            num_cols = len(s["df"].columns)
                            
                            ws_officer.page_setup.orientation = "portrait"
                            ws_officer.page_setup.fitToWidth = 1
                            ws_officer.page_setup.fitToHeight = 0
                            ws_officer.sheet_properties.pageSetUpPr.fitToPage = True
                            
                            title_row = ["員警開單績效統計表"] + [None] * (num_cols - 1)
                            if s["print_date"]:
                                put_idx = min(3, num_cols - 1) 
                                title_row[put_idx] = s["print_date"]
                            ws_officer.append(title_row)
                            ws_officer['A1'].font = Font(size=14, bold=True)
                            if s["print_date"]:
                                date_cell = ws_officer.cell(row=1, column=put_idx + 1)
                                date_cell.font = Font(bold=True, color="0000FF") 
                                date_cell.alignment = Alignment(horizontal="left")
                                
                            if s["issue_date"]:
                                ws_officer.append([s["issue_date"]] + [None] * (num_cols - 1))
                                
                            if s["unit_val"]:
                                ws_officer.append([s["unit_val"]] + [None] * (num_cols - 1))
                            else:
                                ws_officer.append([f"舉發單位：{unit_name}"] + [None] * (num_cols - 1))
                                
                            if s["officer_val"]:
                                ws_officer.append([s["officer_val"]] + [None] * (num_cols - 1))
                            else:
                                ws_officer.append([f"舉發員警：{s['officer']}"] + [None] * (num_cols - 1))

                            current_row_offset = ws_officer.max_row 
                            
                            for r_idx, row in s["df"].iterrows():
                                cleaned_row = [val if pd.notna(val) and str(val).strip() != "" else None for val in row.tolist()]
                                ws_officer.append(cleaned_row)
                                
                            for r, c in s["yellow_cells"]:
                                excel_r = current_row_offset + r + 1
                                excel_c = c + 1
                                ws_officer.cell(row=excel_r, column=excel_c).fill = yellow_fill
                                
                            officer_summary = df_summary[df_summary['員警姓名'] == s['officer']].iloc[0]
                            
                            ws_officer.delete_cols(2)
                            
                            footer_start_row = ws_officer.max_row + 2
                            
                            def write_footer(row_offset, title, val, color="000000"):
                                c_title = ws_officer.cell(row=footer_start_row + row_offset, column=2, value=title)
                                c_title.font = Font(bold=True)
                                c_title.alignment = Alignment(horizontal="right") 
                                
                                c_val = ws_officer.cell(row=footer_start_row + row_offset, column=3, value=val)
                                c_val.font = Font(bold=True, color=color)
                                c_val.alignment = Alignment(horizontal="left")
                            
                            if period == "上半年":
                                write_footer(0, "上半年總分：", int(officer_summary["上半年總分"]), "0000FF")
                                write_footer(1, "上半年剩餘分數：", int(officer_summary["上半年剩餘分數"]), "FF0000")
                            else:
                                write_footer(0, "本期原始分數：", int(officer_summary["本期原始分數"]), "000000")
                                write_footer(1, "上半年結轉餘數：", int(officer_summary["上半年結轉餘數"]), "000000")
                                write_footer(2, "下半年總分：", int(officer_summary["下半年總分"]), "FF0000")
                            
                            autofit_columns(ws_officer, min_row=current_row_offset, max_width=65)

                        wb.save(output)
                        all_output_buffers[f"{target_roc_year}年{period}交通執法重點工作舉發績效結算表_{unit_name}.xlsx"] = output

                if all_summaries:
                    st.success("✅ 所有單位的原始報表已全數完成結算！")
                    
                    unit_names = list(all_summaries.keys())
                    tabs_result = st.tabs([f"🏢 {name}" for name in unit_names])
                    
                    for i, name in enumerate(unit_names):
                        with tabs_result[i]:
                            st.subheader(f"📊 {name} - 績效結算總表")
                            st.dataframe(all_summaries[name], use_container_width=True, hide_index=True)
                            
                            dl_filename = f"{target_roc_year}年{period}交通執法重點工作舉發績效結算表_{name}.xlsx"
                            st.download_button(
                                label=f"📥 下載 {name} 完整報表",
                                data=all_output_buffers[dl_filename].getvalue(),
                                file_name=dl_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True,
                                key=f"dl_{name}"
                            )

                    st.divider()
                    st.markdown("### 📧 批次寄送報表")
                    st.info("點擊下方按鈕，系統會將上方產出的 **所有單位的 Excel 報表** 一次打包寄送至您的信箱！")
                    
                    if st.button("🚀 一鍵將所有報表寄至我的信箱", use_container_width=True):
                        with st.spinner("信件發送中，請稍候…"):
                            ok, mail_err = send_multiple_files_email(all_output_buffers)
                            if ok:
                                st.success(f"✅ 信件發送成功！本次共寄出 {len(all_output_buffers)} 份報表。")
                                st.balloons()
                            else:
                                st.error(f"❌ 發信失敗: {mail_err}")

# 執行點
if __name__ == "__main__":
    main()
